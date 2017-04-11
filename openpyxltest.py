import openpyxl
wb = openpyxl.load_workbook('rawdata.xlsx')
sheet = wb.active
print(sheet.max_row)
print(sheet.max_column)
'''
wb = Workbook()
ws = wb.active
ws.title = "New Title"
ws['A2']=4
wb.save('openpyxltest.xlsx')
'''